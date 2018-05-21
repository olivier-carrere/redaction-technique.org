#!/usr/bin/python3
# coding: utf8
from openpyxl import load_workbook


def titre(title, ornament):
    print(title)
    for i in title:
        print(ornament, end='')
    print('\n')


def contenu(mincol, maxcol, minrow, maxrow):
    for row in ws.iter_rows(min_col=mincol, max_col=maxcol,
                            min_row=minrow, max_row=maxrow):
        for cell in row:
            print('- ' + cell.value)
    print()

# Charger le fichier Excel en lecture seule
wb = load_workbook(filename='produits.xlsx', read_only=True)
# Extraire le classeur actif
ws = wb.active

# Afficher le contenu ReST à l'écran
titre('Produits et versions', '=')

titre(ws['A1'].value, '-')
contenu(1, 1, 2, 4)

titre(ws['B1'].value, '-')
contenu(2, 2, 2, 4)

titre(ws['C1'].value, '-')
contenu(3, 3, 2, 4)
